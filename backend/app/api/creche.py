"""Reembolso-Creche (IN SEGES/MGI 147/2026): página de acompanhamento do RH.

Nesta 1ª onda entrega o LEVANTAMENTO de elegibilidade por posto — a resposta
que os ofícios (CNMP nº 5/2026, ANATEL nº 45/2026) cobram em 5 dias úteis:
quantos colaboradores estão alocados em postos abrangidos pela IN. A camada de
dados das crianças (idade em anos/meses, documentos) entra na 2ª onda, quando o
autocadastro público estiver no ar; a estrutura já a comporta."""

import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, Response
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.auth_rh import requer_rh
from app.core.db import get_db
from app.models.candidato import Candidato, PostoServico
from app.models.usuario_rh import UsuarioRH
from app.services.auditoria import registrar

router = APIRouter(tags=["creche-rh"], dependencies=[Depends(requer_rh)])


def _postos_elegiveis(db: Session) -> list[PostoServico]:
    return db.scalars(
        select(PostoServico)
        .where(PostoServico.da_direito_creche == True)  # noqa: E712
        .order_by(PostoServico.nome)
    ).all()


@router.get("/rh/creche/resumo")
def resumo(db: Session = Depends(get_db)) -> dict:
    """Panorama do benefício: total de postos elegíveis e de colaboradores
    ativos alocados neles, quebrado por posto (com o valor de cada contrato)."""
    postos = _postos_elegiveis(db)
    ids = [p.id for p in postos]
    # contagem de colaboradores ATIVOS por posto elegível (uma consulta só)
    por_posto: dict = {pid: 0 for pid in ids}
    if ids:
        ativos = db.scalars(
            select(Candidato).where(
                Candidato.posto_servico_id.in_(ids),
                Candidato.situacao == "ativo",
            )
        ).all()
        for c in ativos:
            por_posto[c.posto_servico_id] = por_posto.get(c.posto_servico_id, 0) + 1

    linhas = [{
        "posto_id": p.id, "posto": p.nome, "sigla": p.sigla,
        "contrato_ref": p.contrato_ref,
        "valor_reembolso": p.valor_reembolso_creche,
        "colaboradores_ativos": por_posto.get(p.id, 0),
    } for p in postos]

    return {
        "postos_elegiveis": len(postos),
        "colaboradores_em_postos_elegiveis": sum(por_posto.values()),
        "por_posto": linhas,
    }


@router.get("/rh/creche/exportar")
def exportar(db: Session = Depends(get_db),
             rh: UsuarioRH = Depends(requer_rh)) -> Response:
    """Excel do levantamento: um colaborador ativo por linha, em postos que dão
    direito ao benefício, com o valor do reembolso do contrato. É a relação
    nominal que os órgãos pedem para instruir a repactuação."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    postos = {p.id: p for p in _postos_elegiveis(db)}
    colaboradores = []
    if postos:
        colaboradores = db.scalars(
            select(Candidato).where(
                Candidato.posto_servico_id.in_(list(postos.keys())),
                Candidato.situacao == "ativo",
            ).order_by(Candidato.nome_completo)
        ).all()

    cols = ["Nome completo", "CPF", "Matrícula", "Posto (contrato)", "Sigla",
            "Nº do contrato", "Valor do reembolso", "Data de admissão"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Elegiveis Reembolso-Creche"
    verde = PatternFill("solid", fgColor="0FB257")
    for j, nome in enumerate(cols, start=1):
        cel = ws.cell(row=1, column=j, value=nome)
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = verde
        cel.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(j)].width = max(14, min(40, len(nome) + 8))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    for i, c in enumerate(colaboradores, start=2):
        p = postos.get(c.posto_servico_id)
        valores = [c.nome_completo, c.cpf, c.matricula,
                   p.nome if p else "", p.sigla if p else "",
                   p.contrato_ref if p else "",
                   p.valor_reembolso_creche if p else "", c.data_admissao]
        for j, v in enumerate(valores, start=1):
            ws.cell(row=i, column=j, value=v or "")

    buf = io.BytesIO()
    wb.save(buf)
    registrar(db, "creche_levantamento_exportado", ator="rh", ator_detalhe=rh.email,
              detalhe={"colaboradores": len(colaboradores), "postos": len(postos)})
    db.commit()
    agora = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="reembolso-creche-elegiveis-{agora}.xlsx"'},
    )
